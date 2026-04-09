"""
율이공방 — 예산 관리 데이터 관리 모듈
데스크톱(#006)과 웹앱(#012) 공통 사용
"""

import os, re
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

DEFAULT_YEAR  = 2026
DEFAULT_TOTAL = 120_320_000

DEFAULT_PROJECTS = ["토요상설공연", "무등울림축제", "추모제", "국가유산포럼", "관 공통(부대경비)"]
COMMON_PROJECT   = "관 공통(부대경비)"

DEFAULT_PLAN = [
    ("행사운영비", "201-03", [
        ("홍보물 제작",    3_000_000),
        ("음향임차",       15_000_000),
        ("전통체험 용역",  5_000_000),
        ("사회자 운영",    8_000_000),
        ("다과비",         2_000_000),
        ("정월대보름",     3_000_000),
        ("사무용품",       1_000_000),
    ]),
    ("행사실비보상금", "301-10", [
        ("공연 출연료",    75_000_000),
        ("행사진행인력",   8_320_000),
    ]),
]


def clean_num(s):
    s = str(s).replace(",", "").strip()
    try:
        return int(float(s))
    except Exception:
        return 0


def fmt_won(v, short=False):
    try:
        v = int(v)
    except Exception:
        return str(v)
    if short:
        return f"{v:,}천원"
    return f"{v:,}원 ({v//1000:,}천원)"


class ProjectData:
    """사업 1개의 데이터"""
    def __init__(self):
        self.year         = DEFAULT_YEAR
        self.total_budget = 0
        self.categories   = []   # [{name, code, budget, items:[{name,budget}]}]
        self.monthly      = {}   # {cat_name: {1:금액, ..., 12:금액}}
        self.records      = []   # [{id,date,cat,item,detail,amount,round_,memo}]
        self._next_id     = 1


class DataManager:
    def __init__(self, db_file):
        self.db_file = db_file
        self.project_names = []
        self.projects      = {}
        self.current_project = ""
        self.load()

    # ── 현재 사업 프로퍼티 (호환) ──
    @property
    def _p(self):
        return self.projects.get(self.current_project) or ProjectData()

    @property
    def year(self):         return self._p.year
    @year.setter
    def year(self, v):      self._p.year = v

    @property
    def total_budget(self): return self._p.total_budget
    @total_budget.setter
    def total_budget(self, v): self._p.total_budget = v

    @property
    def categories(self):   return self._p.categories
    @categories.setter
    def categories(self, v): self._p.categories = v

    @property
    def monthly(self):      return self._p.monthly
    @monthly.setter
    def monthly(self, v):   self._p.monthly = v

    @property
    def records(self):      return self._p.records
    @records.setter
    def records(self, v):   self._p.records = v

    @property
    def _next_id(self):     return self._p._next_id
    @_next_id.setter
    def _next_id(self, v):  self._p._next_id = v

    @property
    def is_common(self):
        return self.current_project == COMMON_PROJECT

    def switch_project(self, name):
        if name in self.projects:
            self.current_project = name

    def add_project(self, name):
        if name and name not in self.projects:
            self.project_names.append(name)
            self.projects[name] = ProjectData()
            self.current_project = name

    def rename_project(self, old_name, new_name):
        if old_name not in self.projects or not new_name:
            return False
        if new_name in self.projects:
            return False
        idx = self.project_names.index(old_name)
        self.project_names[idx] = new_name
        self.projects[new_name] = self.projects.pop(old_name)
        if self.current_project == old_name:
            self.current_project = new_name
        return True

    def delete_project(self, name):
        if name in self.projects:
            del self.projects[name]
            self.project_names = [n for n in self.project_names if n != name]
            self.current_project = self.project_names[0] if self.project_names else ""

    # ── 불러오기 ──
    def load(self):
        if not os.path.exists(self.db_file):
            self._init_default()
            self.save()
            return
        try:
            xl = pd.ExcelFile(self.db_file, engine="openpyxl")
            sheets = xl.sheet_names

            if "사업목록" in sheets:
                df_pj = xl.parse("사업목록").fillna("")
                for _, row in df_pj.iterrows():
                    pn = str(row.get("사업명", "")).strip()
                    if pn and pn != "nan":
                        self.project_names.append(pn)
                        self.projects[pn] = ProjectData()

            for pn in list(self.project_names):
                p = self.projects[pn]
                ps = pn.replace("/", "_")
                if f"{ps}_계획" in sheets:
                    self._load_plan(p, xl.parse(f"{ps}_계획"))
                if f"{ps}_집행내역" in sheets:
                    self._load_records(p, xl.parse(f"{ps}_집행내역"))
                if f"{ps}_월별배분" in sheets:
                    self._load_monthly(p, xl.parse(f"{ps}_월별배분"))

            # 마이그레이션: 구 형식
            if not self.project_names and "계획" in sheets:
                pn = "토요상설공연"
                self.project_names.append(pn)
                self.projects[pn] = ProjectData()
                p = self.projects[pn]
                self._load_plan(p, xl.parse("계획"))
                if "집행내역" in sheets:
                    self._load_records(p, xl.parse("집행내역"))
                if "월별배분" in sheets:
                    self._load_monthly(p, xl.parse("월별배분"))
                for dpn in DEFAULT_PROJECTS:
                    if dpn not in self.projects:
                        self.project_names.append(dpn)
                        self.projects[dpn] = ProjectData()

            if not self.project_names:
                self._init_default()

            self.current_project = self.project_names[0] if self.project_names else ""

        except Exception:
            self._init_default()

    def _init_default(self):
        self.project_names = list(DEFAULT_PROJECTS)
        self.projects = {}
        for pn in self.project_names:
            self.projects[pn] = ProjectData()
        p = self.projects["토요상설공연"]
        p.year = DEFAULT_YEAR
        p.total_budget = DEFAULT_TOTAL
        for cname, ccode, items in DEFAULT_PLAN:
            cat = {"name": cname, "code": ccode,
                   "budget": sum(i[1] for i in items),
                   "items": [{"name": n, "budget": b} for n, b in items]}
            p.categories.append(cat)
        self.current_project = self.project_names[0]

    def _load_plan(self, p, df):
        df = df.fillna("")
        if not df.empty:
            try:
                p.year = int(df.iloc[0].get("연도", DEFAULT_YEAR) or DEFAULT_YEAR)
                p.total_budget = int(df.iloc[0].get("총예산", 0) or 0)
            except Exception:
                pass
        p.categories = []
        cat_map = {}
        for _, row in df.iterrows():
            cn = str(row.get("편성목명", "")).strip()
            cc = str(row.get("편성목코드", "")).strip()
            cb = clean_num(row.get("편성목예산", 0))
            iname = str(row.get("세부항목명", "")).strip()
            ibud  = clean_num(row.get("세부항목예산", 0))
            if cn and cn not in cat_map:
                cat = {"name": cn, "code": cc, "budget": cb, "items": []}
                p.categories.append(cat)
                cat_map[cn] = cat
            if cn in cat_map and iname:
                cat_map[cn]["items"].append({"name": iname, "budget": ibud})

    def _load_records(self, p, df):
        df = df.fillna("")
        p.records = []
        p._next_id = 1
        for _, row in df.iterrows():
            rid = str(row.get("ID", "")).strip()
            if not rid or rid == "nan":
                continue
            try:
                rid_int = int(rid)
                if rid_int >= p._next_id:
                    p._next_id = rid_int + 1
            except Exception:
                pass
            p.records.append({
                "id":      rid,
                "date":    str(row.get("집행일", "")).strip(),
                "cat":     str(row.get("편성목", "")).strip(),
                "item":    str(row.get("세부항목", "")).strip(),
                "detail":  str(row.get("세부내용", "")).strip(),
                "amount":  clean_num(row.get("금액", 0)),
                "round_":  str(row.get("회차", "")).strip(),
                "memo":    str(row.get("비고", "")).strip(),
            })

    def _load_monthly(self, p, df):
        df = df.fillna(0)
        p.monthly = {}
        for _, row in df.iterrows():
            cn = str(row.get("편성목명", "")).strip()
            if not cn or cn == "nan":
                continue
            p.monthly[cn] = {m: clean_num(row.get(str(m), 0)) for m in range(1, 13)}

    # ── 저장 ──
    def save(self):
        wb = openpyxl.Workbook()

        ws_pj = wb.active; ws_pj.title = "사업목록"
        ws_pj.append(["사업명"]); self._hdr(ws_pj, 1, 1)
        for pn in self.project_names:
            ws_pj.append([pn])

        for pn in self.project_names:
            p = self.projects[pn]
            ps = pn.replace("/", "_")
            ws1 = wb.create_sheet(f"{ps}_계획")
            self._save_plan(ws1, p)
            ws2 = wb.create_sheet(f"{ps}_집행내역")
            self._save_records(ws2, p)
            ws3 = wb.create_sheet(f"{ps}_월별배분")
            self._save_monthly(ws3, p)

        wb.save(self.db_file)

    def _hdr(self, ws, row, ncols):
        fill = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
        font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center")

    def _save_plan(self, ws, p):
        cols = ["연도","총예산","편성목명","편성목코드","편성목예산","세부항목명","세부항목예산"]
        ws.append(cols); self._hdr(ws, 1, len(cols))
        first = True
        for cat in p.categories:
            for item in cat["items"]:
                ws.append([
                    p.year if first else "", p.total_budget if first else "",
                    cat["name"], cat["code"], cat["budget"],
                    item["name"], item["budget"]
                ])
                first = False
            if not cat["items"]:
                ws.append([p.year if first else "", p.total_budget if first else "",
                           cat["name"], cat["code"], cat["budget"], "", ""])
                first = False

    def _save_records(self, ws, p):
        cols = ["ID","집행일","편성목","세부항목","세부내용","금액","회차","비고"]
        ws.append(cols); self._hdr(ws, 1, len(cols))
        for r in p.records:
            ws.append([r["id"],r["date"],r["cat"],r["item"],
                       r["detail"],r["amount"],r["round_"],r["memo"]])

    def _save_monthly(self, ws, p):
        cols = ["편성목명"] + [str(m) for m in range(1, 13)]
        ws.append(cols); self._hdr(ws, 1, len(cols))
        for cat in p.categories:
            md = p.monthly.get(cat["name"], {})
            ws.append([cat["name"]] + [md.get(m, 0) for m in range(1, 13)])

    # ── 계산 헬퍼 ──
    def cat_spent(self, cat_name):
        return sum(r["amount"] for r in self.records if r["cat"] == cat_name)

    def item_spent(self, cat_name, item_name):
        return sum(r["amount"] for r in self.records
                   if r["cat"] == cat_name and r["item"] == item_name)

    def monthly_spent(self, month):
        return sum(r["amount"] for r in self.records
                   if r["date"] and r["date"][5:7] == f"{month:02d}")

    def monthly_plan(self, month):
        return sum(md.get(month, 0) for md in self.monthly.values())

    def total_spent(self):
        return sum(r["amount"] for r in self.records)

    def get_cat(self, name):
        for c in self.categories:
            if c["name"] == name:
                return c
        return None

    def get_item(self, cat_name, item_name):
        cat = self.get_cat(cat_name)
        if cat:
            for it in cat["items"]:
                if it["name"] == item_name:
                    return it
        return None

    def sync_cat_budget(self, cat_name):
        cat = self.get_cat(cat_name)
        if cat and cat["items"]:
            cat["budget"] = sum(it["budget"] for it in cat["items"])

    def all_project_cat_names(self):
        """전체 사업의 편성목 이름 수집 (관 공통용)"""
        names = set()
        for pn in self.project_names:
            p = self.projects[pn]
            for c in p.categories:
                names.add(c["name"])
        used = set(r["cat"] for r in self.records)
        return sorted(names | used)

    # ── 집행 CRUD ──
    def add_record(self, data):
        data["id"] = str(self._next_id)
        self._next_id += 1
        self.records.append(data)
        self.save()

    def update_record(self, rid, data):
        for i, r in enumerate(self.records):
            if r["id"] == rid:
                data["id"] = rid
                self.records[i] = data
                break
        self.save()

    def delete_record(self, rid):
        self.records = [r for r in self.records if r["id"] != rid]
        self.save()

    # ── 엑셀 내보내기 헬퍼 ──
    def export_settlement_wb(self):
        """정산표 워크북 생성"""
        wb = openpyxl.Workbook()
        bold_font = Font(name="맑은 고딕", bold=True)
        num_fmt = "#,##0"

        ws1 = wb.active; ws1.title = "편성목별 정산"
        cols1 = ["편성목", "세부항목", "세부내용", "집행일", "금액(원)", "회차", "비고"]
        ws1.append(cols1); self._hdr(ws1, 1, len(cols1))
        sorted_recs = sorted(self.records,
            key=lambda r: (r["cat"], r["item"], r["date"]))
        for r in sorted_recs:
            ws1.append([r["cat"], r["item"], r["detail"], r["date"],
                        r["amount"], r["round_"], r["memo"]])
            ws1.cell(ws1.max_row, 5).number_format = num_fmt
        total_row = ws1.max_row + 1
        ws1.cell(total_row, 4, "합  계").font = bold_font
        ws1.cell(total_row, 5, sum(r["amount"] for r in self.records)).font = bold_font
        ws1.cell(total_row, 5).number_format = num_fmt

        ws2 = wb.create_sheet("편성목별 요약")
        cols2 = ["편성목", "코드", "예산(원)", "집행액(원)", "잔액(원)", "집행률(%)"]
        ws2.append(cols2); self._hdr(ws2, 1, len(cols2))
        for cat in self.categories:
            sp = self.cat_spent(cat["name"])
            rem = cat["budget"] - sp
            rate = round(sp/cat["budget"]*100, 1) if cat["budget"] else 0
            ws2.append([cat["name"], cat["code"], cat["budget"], sp, rem, rate])
            for ci in [3, 4, 5]:
                ws2.cell(ws2.max_row, ci).number_format = num_fmt

        ws3 = wb.create_sheet("세부항목별 요약")
        cols3 = ["편성목", "세부항목", "예산(원)", "집행액(원)", "잔액(원)", "집행률(%)"]
        ws3.append(cols3); self._hdr(ws3, 1, len(cols3))
        for cat in self.categories:
            for item in cat["items"]:
                sp = self.item_spent(cat["name"], item["name"])
                rem = item["budget"] - sp
                rate = round(sp/item["budget"]*100, 1) if item["budget"] else 0
                ws3.append([cat["name"], item["name"], item["budget"], sp, rem, rate])
                for ci in [3, 4, 5]:
                    ws3.cell(ws3.max_row, ci).number_format = num_fmt

        if self.is_common and self.records:
            ws4 = wb.create_sheet("비고별 집행 현황")
            cols4 = ["비고(관련사업)", "집행 건수", "집행 합계(원)"]
            ws4.append(cols4); self._hdr(ws4, 1, len(cols4))
            memo_data = {}
            for r in self.records:
                memo = r.get("memo", "").strip() or "(미분류)"
                if memo not in memo_data:
                    memo_data[memo] = {"count": 0, "amount": 0}
                memo_data[memo]["count"] += 1
                memo_data[memo]["amount"] += r["amount"]
            for memo, d in sorted(memo_data.items(), key=lambda x: -x[1]["amount"]):
                ws4.append([memo, d["count"], d["amount"]])
                ws4.cell(ws4.max_row, 3).number_format = num_fmt
            total_row = ws4.max_row + 1
            ws4.cell(total_row, 1, "합  계").font = bold_font
            ws4.cell(total_row, 2, sum(d["count"] for d in memo_data.values())).font = bold_font
            ws4.cell(total_row, 3, sum(d["amount"] for d in memo_data.values())).font = bold_font
            ws4.cell(total_row, 3).number_format = num_fmt

        return wb

    def export_monthly_wb(self):
        """월별 현황 워크북 생성"""
        wb = openpyxl.Workbook()
        num_fmt = "#,##0"

        ws1 = wb.active; ws1.title = "월별 요약"
        cols1 = ["월", "계획(원)", "집행(원)", "차이(원)", "달성률(%)"]
        ws1.append(cols1); self._hdr(ws1, 1, len(cols1))
        for m in range(1, 13):
            plan = self.monthly_plan(m)
            sp   = self.monthly_spent(m)
            diff = sp - plan
            rate = round(sp/plan*100, 1) if plan else None
            ws1.append([f"{m}월", plan if plan else None, sp,
                        diff if plan else None, rate])
            for ci in [2, 3, 4]:
                ws1.cell(ws1.max_row, ci).number_format = num_fmt

        ws2 = wb.create_sheet("월별×편성목")
        cat_names = [c["name"] for c in self.categories]
        ws2.append(["월"] + cat_names + ["합계"])
        self._hdr(ws2, 1, len(cat_names)+2)
        for m in range(1, 13):
            row_vals = [f"{m}월"]
            row_total = 0
            for cn in cat_names:
                sp = sum(r["amount"] for r in self.records
                         if r["cat"] == cn and r["date"][5:7] == f"{m:02d}")
                row_vals.append(sp); row_total += sp
            row_vals.append(row_total)
            ws2.append(row_vals)
            for ci in range(2, len(cat_names)+3):
                ws2.cell(ws2.max_row, ci).number_format = num_fmt

        return wb
